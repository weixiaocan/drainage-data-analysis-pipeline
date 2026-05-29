 # -*- coding: utf-8 -*-
###
#读取原始数据并统计数据量
###
##导入变量
import pandas as pd
import numpy as np
import pickle
from openpyxl import Workbook,load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#读取基本信息
from base_info import folder_name,equip_name,node_num,node_name,flow_header_names

## 数据读取
def data_read(folder_name,node_num,equip_name,node_name):
    #生成数据读取的路径
    path = []
    for i in range(node_num):
        path.append('../'+folder_name+'/' + equip_name[i] + '.csv')
    #原始数据的读取和储存
    flow_data_ori = {}
    for i in range(node_num):
        flow_data_ori[node_name[i]] = pd.read_csv(path[i],skiprows=[0],header = None ,usecols=[0,3,4,5],names=flow_header_names,index_col = 0,parse_dates = [0]
                                                    ,dtype = {'l':np.float64,'f':np.float64,'velo':np.float64})
        
    return flow_data_ori

##原始数据量统计
def num_sta(flow_data_ori):
    Num_of_data_true = np.empty(node_num) #每个点位监测到的数据条数
    Num_of_day = np.empty(node_num) #监测天数
    Num_of_data_exp = np.empty(node_num) #理论监测数据条数
    Miss_rate = np.empty(node_num) #数据缺失率
    for i in range(node_num):
        Num_of_data_true[i] = flow_data_ori[node_name[i]].shape[0]        
        Num_of_day[i] = (flow_data_ori[node_name[i]].index[-1] - flow_data_ori[node_name[i]].index[0]).days + 1
    #       Num_of_data_exp[i] = Num_of_day[i]*24*60
        Num_of_data_exp[i] = (flow_data_ori[node_name[i]].index[-1] - flow_data_ori[node_name[i]].index[0]).total_seconds()/60+1
        Miss_rate[i] = 1-Num_of_data_true[i]/Num_of_data_exp[i]
    #生成DF
    data_num_sta = pd.DataFrame({'Num_of_data_true':Num_of_data_true,'Num_of_day':Num_of_day,
                                 'Num_of_data_exp':Num_of_data_exp,'Miss_rate':Miss_rate},index=node_name)
    return data_num_sta

##保存统计结果至EXCEL表格
def save_to_excel(data,excel_name,sheet_name,hds):
    # 打开或新建统计表格
    try:
        wb = load_workbook(excel_name)
    except FileNotFoundError:
        wb = Workbook()
    # 新建表单
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    ws = wb.create_sheet(sheet_name)
    # 数据写入
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r) 
    # 删除第二行
    ws.delete_rows(2)
    # 替换表头
    for i in range(len(hds)):
        ws.cell(1,i+1).value=hds[i]
    # 存储表格
    wb.save(excel_name)

if __name__ == '__main__':  
    ## 数据读取
    flow_data_ori = data_read(folder_name,node_num,equip_name,node_name)
    #存储原始数据为pickle格式，数据量较大时可减少读取时间
    with open('flow_data_ori.pickle', 'wb') as f:
        # Pickle the 'data' dictionary using the highest protocol available.
        pickle.dump(flow_data_ori, f, pickle.HIGHEST_PROTOCOL)
    #读取格式
    #with open('flow_data_ori.pickle', 'rb') as f:
    #    flow_data_ori = pickle.load(f)
    ##原始数据量统计
    data_num_sta = num_sta(flow_data_ori)
    ##保存统计结果至EXCEL表格    
    excel_name = '../statistics.xlsx'
    sheet_name = '数据总量统计'
    hds=['点位编号','监测数据条数','监测天数','理论数据条数','数据缺失率']
    save_to_excel(data_num_sta,excel_name,sheet_name,hds)      
