"""雨天事件统计核心逻辑

统计降雨事件下各点位的基本数据：
- 最大液位
- 平均流量
- 峰值流量等
"""

from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class EventStatsConfig:
    """雨天事件统计配置参数"""
    rain_effect_delay: float = 12.0  # 降雨效应延迟时间（小时）


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV"""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as err:
            last_err = err
    if last_err:
        raise last_err
    raise RuntimeError(f"无法读取 CSV: {path}")


def _detect_columns(df: pd.DataFrame) -> tuple[str, str, str, str | None]:
    """检测 CSV 列名"""
    cols = [str(c).strip() for c in df.columns]
    time_col = "数据时间" if "数据时间" in cols else next(c for c in cols if "时间" in c)
    flow_col = "流量(L/s)(均值)" if "流量(L/s)(均值)" in cols else next(c for c in cols if "流量" in c)
    level_col = "液位(m)(均值)" if "液位(m)(均值)" in cols else next(c for c in cols if "液位" in c)
    velocity_col = None
    for c in cols:
        if "流速" in c:
            velocity_col = c
            break
    return time_col, flow_col, level_col, velocity_col


def _parse_point_name(path: Path) -> str:
    """从文件名解析点位编号"""
    stem = path.stem
    if "_" in stem:
        return stem.split("_", 1)[1]
    return stem


def _load_flow_data(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """加载流量数据"""
    result: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        df = _read_csv_with_fallback(csv_path)
        time_col, flow_col, level_col, velocity_col = _detect_columns(df)

        rename_map = {time_col: "数据时间", flow_col: "f", level_col: "l"}
        if velocity_col:
            rename_map[velocity_col] = "velo"
        df = df.rename(columns=rename_map)

        df["数据时间"] = pd.to_datetime(df["数据时间"], errors="coerce")
        df = df.dropna(subset=["数据时间"]).copy()
        df["f"] = pd.to_numeric(df["f"], errors="coerce").fillna(0.0)
        df["l"] = pd.to_numeric(df["l"], errors="coerce").fillna(0.0)

        point_name = _parse_point_name(csv_path)
        result[point_name] = df.sort_values("数据时间").reset_index(drop=True)
    return result


def _get_event_stats(
    flow_data: dict[str, pd.DataFrame],
    event_data: dict[int, dict],
    delay_hours: float,
    selected_events: list[int] | None = None,
) -> pd.DataFrame:
    """统计降雨事件下各点位的数据

    Args:
        flow_data: 流量数据字典
        event_data: 场次降雨数据
        delay_hours: 降雨效应延迟时间
        selected_events: 选中的场次编号列表

    Returns:
        统计结果 DataFrame
    """
    point_names = list(flow_data.keys())
    event_ids = sorted(event_data.keys())

    # 过滤选中的场次
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    rows: list[dict] = []

    for point_name in point_names:
        df = flow_data[point_name]
        row: dict[str, Any] = {"点位编号": point_name}

        for event_id in event_ids:
            event = event_data[event_id]
            start = event["start"]
            end = event["end"] + timedelta(hours=delay_hours)

            event_df = df[(df["数据时间"] >= start) & (df["数据时间"] <= end)]

            if len(event_df) > 0:
                # 最大液位
                max_level = event_df["l"].max()
                # 平均流量 (m³/d)
                avg_flow = event_df["f"].mean() * 86.4
                # 峰值流量 (L/s)
                peak_flow = event_df["f"].max()
                # 流量标准差
                flow_std = event_df["f"].std()

                row[f"场次{event_id}_最大液位(m)"] = round(max_level, 2)
                row[f"场次{event_id}_平均流量(m³/d)"] = round(avg_flow, 2)
                row[f"场次{event_id}_峰值流量(L/s)"] = round(peak_flow, 2)
                row[f"场次{event_id}_流量标准差"] = round(flow_std, 2)
            else:
                row[f"场次{event_id}_最大液位(m)"] = np.nan
                row[f"场次{event_id}_平均流量(m³/d)"] = np.nan
                row[f"场次{event_id}_峰值流量(L/s)"] = np.nan
                row[f"场次{event_id}_流量标准差"] = np.nan

        rows.append(row)

    return pd.DataFrame(rows)


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_event_stats(
    flow_dir: Path,
    event_data: dict[int, dict],
    combined_xlsx: Path,
    selected_events: list[int] | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行雨天事件统计

    Args:
        flow_dir: 流量数据目录
        event_data: 场次降雨数据
        combined_xlsx: 综合分析结果 xlsx 文件
        selected_events: 选中的场次编号列表
        config: 可选配置参数

    Returns:
        {
            "event_stats": pd.DataFrame,
        }
    """
    # 合并配置
    cfg = EventStatsConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载流量数据
    print(f"读取流量数据: {flow_dir}")
    flow_data = _load_flow_data(flow_dir)
    print(f"  - 点位数: {len(flow_data)}")

    print(f"使用场次降雨数据: {len(event_data)} 个场次")
    if selected_events:
        print(f"  - 选中场次: {selected_events}")

    # 统计降雨事件
    print(f"统计降雨事件数据 (延迟时间: {cfg.rain_effect_delay}小时)")
    event_stats_df = _get_event_stats(
        flow_data, event_data, cfg.rain_effect_delay, selected_events
    )

    # 保存到 Excel
    # 构建表头
    event_ids = selected_events if selected_events else sorted(event_data.keys())
    headers = ["点位编号"]
    for event_id in event_ids:
        headers.extend([
            f"场次{event_id}_最大液位(m)",
            f"场次{event_id}_平均流量(m³/d)",
            f"场次{event_id}_峰值流量(L/s)",
            f"场次{event_id}_流量标准差",
        ])

    _save_to_excel(event_stats_df, combined_xlsx, "雨天事件统计", headers)
    print(f"保存雨天事件统计: {combined_xlsx}")

    return {
        "event_stats": event_stats_df,
    }
