"""RDII分析核心逻辑

从 colleague_tool/code/analyze_event_flow.py 和 analyze_event_RDII.py 提取核心逻辑。

RDII (Rainfall-Derived Infiltration and Inflow) = 雨天流量 - 旱天特征曲线
"""

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class RDIIConfig:
    """RDII分析配置参数"""
    rain_effect_delay: float = 12.0  # 降雨效应延迟时间（小时）


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV"""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as err:
            last_err = err
    if last_err:
        raise last_err
    raise RuntimeError(f"无法读取 CSV: {path}")


def _detect_columns(df: pd.DataFrame) -> tuple[str, str, str, str | None]:
    """检测 CSV 列名"""
    cols = [str(c).strip() for c in df.columns]
    time_col = "数据时间" if "数据时间" in cols else next(c for c in cols if "时间" in c)
    flow_col = "流量(L/s)(均值)" if "流量(L/s)(均值)" in cols else next(c for c in cols if "流量" in c)
    level_col = "液位(m)(均值)" if "液位(m)(均值)" in cols else next(c for c in cols if "液位" in c)
    velocity_col = None
    for c in cols:
        if "流速" in c:
            velocity_col = c
            break
    return time_col, flow_col, level_col, velocity_col


def _parse_point_name(path: Path) -> str:
    """从文件名解析点位编号"""
    stem = path.stem
    if "_" in stem:
        return stem.split("_", 1)[1]
    return stem


def _load_flow_data(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """加载流量数据"""
    result: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        df = _read_csv_with_fallback(csv_path)
        time_col, flow_col, level_col, velocity_col = _detect_columns(df)

        rename_map = {time_col: "数据时间", flow_col: "f", level_col: "l"}
        if velocity_col:
            rename_map[velocity_col] = "velo"
        df = df.rename(columns=rename_map)

        df["数据时间"] = pd.to_datetime(df["数据时间"], errors="coerce")
        df = df.dropna(subset=["数据时间"]).copy()
        df["f"] = pd.to_numeric(df["f"], errors="coerce").fillna(0.0)
        df["l"] = pd.to_numeric(df["l"], errors="coerce").fillna(0.0)

        point_name = _parse_point_name(csv_path)
        result[point_name] = df.sort_values("数据时间").reset_index(drop=True)
    return result


def _get_event_flow_stats(
    flow_data: dict[str, pd.DataFrame],
    event_data: dict,
    delay_hours: float,
    selected_events: list[int] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """统计降雨事件下各点位的最大液位和平均流量

    Args:
        flow_data: 流量数据字典
        event_data: 场次降雨数据（来自 rainfall_analysis）
        delay_hours: 降雨效应延迟时间
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）

    Returns:
        (最大液位DataFrame, 平均流量DataFrame)
    """
    point_names = list(flow_data.keys())
    event_ids = sorted(event_data.keys())

    # 过滤选中的场次
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    max_level_data: dict[str, list] = {"点位编号": point_names}
    avg_flow_data: dict[str, list] = {"点位编号": point_names}

    for event_id in event_ids:
        event = event_data[event_id]
        start = event["start"]
        end = event["end"] + timedelta(hours=delay_hours)

        max_levels = []
        avg_flows = []

        for point_name in point_names:
            df = flow_data[point_name]
            event_df = df[(df["数据时间"] >= start) & (df["数据时间"] <= end)]

            if len(event_df) > 0:
                max_level = event_df["l"].max()
                avg_flow = event_df["f"].mean() * 86.4  # L/s -> m³/d
            else:
                max_level = np.nan
                avg_flow = np.nan

            max_levels.append(round(max_level, 2) if not np.isnan(max_level) else np.nan)
            avg_flows.append(round(avg_flow, 2) if not np.isnan(avg_flow) else np.nan)

        max_level_data[f"场次{event_id}"] = max_levels
        avg_flow_data[f"场次{event_id}"] = avg_flows

    return pd.DataFrame(max_level_data), pd.DataFrame(avg_flow_data)


def _get_rdii_stats(
    flow_data: dict[str, pd.DataFrame],
    dry_curve_data: dict[str, pd.DataFrame],
    event_data: dict,
    delay_hours: float,
    selected_events: list[int] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """计算RDII统计

    Args:
        flow_data: 流量数据字典
        dry_curve_data: 旱天特征曲线
        event_data: 场次降雨数据
        delay_hours: 降雨效应延迟时间
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）

    Returns:
        (RDII总量DataFrame, 雨天流量总量DataFrame, RDII曲线数据字典)
    """
    point_names = list(dry_curve_data.keys())
    event_ids = sorted(event_data.keys())

    # 过滤选中的场次
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    rdii_data: dict[str, list] = {"点位编号": point_names}
    overflow_data: dict[str, list] = {"点位编号": point_names}
    rdii_curve_all: dict[int, dict[str, pd.DataFrame]] = {}

    for event_id in event_ids:
        event = event_data[event_id]
        start = event["start"]
        end = event["end"] + timedelta(hours=delay_hours)

        # 生成时间序列
        delta_minutes = int((end - start).total_seconds() / 60) + 1
        date_index = pd.date_range(start, end, freq="T")

        rdii_values = []
        overflow_values = []
        rdii_curve_event: dict[str, pd.DataFrame] = {}

        for point_name in point_names:
            df = flow_data.get(point_name)
            dry_curve = dry_curve_data.get(point_name)

            if df is None or dry_curve is None:
                rdii_values.append(np.nan)
                overflow_values.append(np.nan)
                continue

            # 获取雨天流量数据
            event_df = df[(df["数据时间"] >= start) & (df["数据时间"] <= end)]
            if len(event_df) != delta_minutes:
                rdii_values.append(np.nan)
                overflow_values.append(np.nan)
                continue

            # 获取对应的旱天特征曲线
            # 从开始时间计算偏移量
            start_minute = start.hour * 60 + start.minute
            dry_flow = dry_curve["f"].values
            dry_flow_tiled = np.tile(dry_flow, int(np.ceil(delta_minutes / 1440)) + 1)
            dry_flow_segment = dry_flow_tiled[start_minute:start_minute + delta_minutes]

            # 计算RDII
            rain_flow = event_df["f"].values
            rdii = rain_flow - dry_flow_segment[:len(rain_flow)]

            # RDII总量 (m³)
            rdii_total = rdii[rdii > 0].sum() * 60 / 1000
            rdii_values.append(round(rdii_total, 2))

            # 雨天流量总量 (m³)
            overflow_total = rain_flow.sum() * 60 / 1000
            overflow_values.append(round(overflow_total, 2))

            # 保存RDII曲线数据
            rdii_curve_event[point_name] = pd.DataFrame({
                "时间": date_index[:len(rain_flow)],
                "雨天流量": rain_flow,
                "旱天流量": dry_flow_segment[:len(rain_flow)],
                "RDII": rdii,
            }).set_index("时间")

        rdii_data[f"场次{event_id}"] = rdii_values
        overflow_data[f"场次{event_id}"] = overflow_values
        rdii_curve_all[event_id] = rdii_curve_event

    return pd.DataFrame(rdii_data), pd.DataFrame(overflow_data), rdii_curve_all


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_rdii_analysis(
    flow_dir: Path,
    dry_curve_data: dict[str, pd.DataFrame],
    event_data: dict[int, dict],
    combined_xlsx: Path,
    selected_events: list[int] | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行RDII分析

    Args:
        flow_dir: 流量数据目录
        dry_curve_data: 旱天特征曲线数据（从内存传入）
        event_data: 场次降雨数据（从内存传入）
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）
        config: 可选配置参数

    Returns:
        {
            "max_level": pd.DataFrame,      # 最大液位统计
            "avg_flow": pd.DataFrame,       # 平均流量统计
            "rdii_total": pd.DataFrame,     # RDII总量统计
            "overflow_total": pd.DataFrame, # 雨天流量总量
            "rdii_curve_data": dict,        # RDII曲线数据
        }
    """
    # 合并配置
    cfg = RDIIConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载流量数据
    print(f"读取流量数据: {flow_dir}")
    flow_data = _load_flow_data(flow_dir)
    print(f"  - 点位数: {len(flow_data)}")

    print(f"使用旱天特征曲线: {len(dry_curve_data)} 个点位")
    print(f"使用场次降雨数据: {len(event_data)} 个场次")

    if selected_events:
        print(f"  - 选中场次: {selected_events}")

    # 统计降雨事件下的流量和液位
    print("统计降雨事件下的流量和液位")
    max_level_df, avg_flow_df = _get_event_flow_stats(
        flow_data, event_data, cfg.rain_effect_delay, selected_events
    )

    # 计算RDII
    print(f"计算RDII (延迟时间: {cfg.rain_effect_delay}小时)")
    rdii_total_df, overflow_total_df, rdii_curve_data = _get_rdii_stats(
        flow_data, dry_curve_data, event_data, cfg.rain_effect_delay, selected_events
    )

    # 保存统计结果到 Excel
    event_ids_used = selected_events if selected_events else sorted(event_data.keys())

    _save_to_excel(
        max_level_df,
        combined_xlsx,
        "降雨事件最大液位",
        ["点位编号"] + [f"场次{e}" for e in event_ids_used]
    )
    print(f"保存最大液位统计: {combined_xlsx}")

    _save_to_excel(
        avg_flow_df,
        combined_xlsx,
        "降雨事件平均流量",
        ["点位编号"] + [f"场次{e}" for e in event_ids_used]
    )
    print(f"保存平均流量统计: {combined_xlsx}")

    _save_to_excel(
        rdii_total_df,
        combined_xlsx,
        "RDII总量统计",
        ["点位编号"] + [f"场次{e}" for e in event_ids_used]
    )
    print(f"保存RDII总量统计: {combined_xlsx}")

    _save_to_excel(
        overflow_total_df,
        combined_xlsx,
        "雨天流量总量",
        ["点位编号"] + [f"场次{e}" for e in event_ids_used]
    )
    print(f"保存雨天流量总量: {combined_xlsx}")

    return {
        "max_level": max_level_df,
        "avg_flow": avg_flow_df,
        "rdii_total": rdii_total_df,
        "overflow_total": overflow_total_df,
        "rdii_curve_data": rdii_curve_data,
    }
