"""排污规律分析核心逻辑

从 _archive_old_agents/agents/pattern_agent.py 和 utils/pattern_feature_engine.py 提取核心逻辑。

基于旱天特征曲线判断排污规律：
- 第1类：符合生活用水规律（早晚高峰明显）
- 第2类：有波峰但不符合典型规律
- 第3类：曲线平坦/异常
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.signal import find_peaks


@dataclass
class PatternConfig:
    """排污规律分析配置参数"""
    kz_life_min: float = 1.3        # 生活污水Kz下限
    kz_life_max: float = 2.5        # 生活污水Kz上限
    kz_flat_threshold: float = 1.2  # 低于此值认为平坦曲线
    peak_prominence_ratio: float = 0.15  # 峰相对于均值的最小突出度
    peak_distance_minutes: int = 120     # 两个峰之间最小间隔
    peak_valley_ratio_threshold: float = 1.5  # 峰谷比低于此值归为第3类
    mean_flow_low_threshold: float = 0.5      # 均值低于此值归为第3类
    night_ratio_high: float = 0.5   # 夜间占比高于此值认为异常


def _calculate_features(curve: pd.DataFrame) -> dict:
    """计算特征值

    Args:
        curve: 特征曲线 DataFrame，包含 'f', 'l', 'velo' 列

    Returns:
        特征字典
    """
    features = {}
    flow = curve["f"].values

    # 基础统计
    features["peak_value"] = float(flow.max())
    features["min_value"] = float(flow.min())
    features["mean_value"] = float(flow.mean())
    features["std_value"] = float(flow.std())

    # 峰谷比
    if features["min_value"] > 0.0001:
        features["peak_valley_ratio"] = features["peak_value"] / features["min_value"]
    else:
        features["peak_valley_ratio"] = 99.9

    # Kz（时变化系数）
    if features["mean_value"] > 0.0001:
        features["kz"] = features["peak_value"] / features["mean_value"]
    else:
        features["kz"] = 0.0

    # 峰识别
    mean_val = features["mean_value"]
    prominence = mean_val * 0.15
    distance = 120  # 分钟

    peaks, properties = find_peaks(flow, prominence=prominence, distance=distance)

    features["peak_count"] = len(peaks)
    if len(peaks) > 0:
        features["peak_times"] = _get_peak_times(curve.index, peaks)
        features["peak_significance"] = float(max(properties["prominences"]) / mean_val) if mean_val > 0 else 0
    else:
        features["peak_times"] = []
        features["peak_significance"] = 0.0

    # 时段分析
    features["morning_peak"] = _analyze_period(flow, curve.index, 6 * 60, 11 * 60)   # 06:00-11:00
    features["evening_peak"] = _analyze_period(flow, curve.index, 18 * 60, 24 * 60)  # 18:00-24:00
    features["night"] = _analyze_period(flow, curve.index, 1 * 60, 5 * 60)           # 01:00-05:00
    features["daytime"] = _analyze_period(flow, curve.index, 6 * 60, 24 * 60)        # 06:00-24:00

    # 夜间/日间比值
    if features["daytime"]["mean"] > 0:
        features["night_day_ratio"] = features["night"]["mean"] / features["daytime"]["mean"]
    else:
        features["night_day_ratio"] = 0.0

    return features


def _get_peak_times(index: pd.DatetimeIndex, peaks: np.ndarray) -> list[str]:
    """获取峰值时间"""
    times = []
    for p in peaks:
        t = index[p]
        times.append(f"{t.hour:02d}:{t.minute:02d}")
    return times


def _analyze_period(values: np.ndarray, index: pd.DatetimeIndex, start_min: int, end_min: int) -> dict:
    """分析指定时段"""
    # 计算每个时间点对应的分钟数
    minutes = np.array([t.hour * 60 + t.minute for t in index])

    if end_min == 24 * 60:
        mask = minutes >= start_min
    else:
        mask = (minutes >= start_min) & (minutes < end_min)

    period_values = values[mask]

    if len(period_values) > 0:
        return {
            "mean": float(period_values.mean()),
            "max": float(period_values.max()),
            "min": float(period_values.min()),
        }
    return {"mean": 0.0, "max": 0.0, "min": 0.0}


def _extract_peak_valley_periods(curve: pd.DataFrame) -> tuple[list[str], list[str]]:
    """提取波峰时段和波谷时段"""
    flow = curve["f"].values
    max_val = flow.max()
    min_val = flow.min()
    threshold = (max_val + min_val) / 2

    # 标记高于/低于阈值
    above = flow > threshold
    below = flow <= threshold

    # 提取连续区间
    peak_periods = _find_continuous_periods(curve.index, above, min_duration=30)
    valley_periods = _find_continuous_periods(curve.index, below, min_duration=30)

    return peak_periods, valley_periods


def _find_continuous_periods(index: pd.DatetimeIndex, mask: np.ndarray, min_duration: int = 30) -> list[str]:
    """提取连续时段"""
    periods = []
    if len(mask) == 0:
        return periods

    in_period = False
    start_idx = 0

    for i, val in enumerate(mask):
        if val and not in_period:
            in_period = True
            start_idx = i
        elif not val and in_period:
            in_period = False
            if i - start_idx >= min_duration:
                start_time = index[start_idx].strftime("%H:%M")
                end_time = index[i - 1].strftime("%H:%M")
                periods.append(f"{start_time}~{end_time}")

    # 处理末尾区间
    if in_period and len(mask) - start_idx >= min_duration:
        start_time = index[start_idx].strftime("%H:%M")
        end_time = index[-1].strftime("%H:%M")
        periods.append(f"{start_time}~{end_time}")

    return periods


def _classify_pattern(features: dict) -> tuple[int, str]:
    """规则分类

    Returns:
        (category, reason)
        category: 1/2/3
    """
    kz = features.get("kz", 0)
    peak_count = features.get("peak_count", 0)
    peak_significance = features.get("peak_significance", 0)
    peak_times = features.get("peak_times", [])
    night_day_ratio = features.get("night_day_ratio", 0)
    peak_valley_ratio = features.get("peak_valley_ratio", 0)
    mean_value = features.get("mean_value", 0)

    # 第3类：流量极低
    if mean_value < 0.5:
        return 3, f"流量接近零(均值={mean_value:.2f}L/s)"

    # 第3类：峰谷比过低
    if peak_valley_ratio < 1.5 and peak_valley_ratio < 99:
        return 3, f"峰谷比={peak_valley_ratio:.2f}(<1.5)，曲线平坦"

    # 第3类：Kz过低或峰不显著
    if kz < 1.2 or peak_significance < 0.3:
        return 3, f"曲线平坦(Kz={kz:.2f}, 峰显著性={peak_significance:.2f})"

    # 检查是否有早晚高峰
    has_morning_peak = any("06:" <= pt <= "10:59" for pt in peak_times)
    has_evening_peak = any("18:" <= pt <= "22:59" for pt in peak_times)

    # 夜间流量是否正常
    night_normal = night_day_ratio < 0.5

    # 第1类：符合生活规律
    if peak_count >= 1 and (has_morning_peak or has_evening_peak) and night_normal:
        if 1.3 <= kz <= 2.5:
            peak_desc = []
            if has_morning_peak:
                peak_desc.append("早高峰")
            if has_evening_peak:
                peak_desc.append("晚高峰")
            return 1, f"符合生活规律(Kz={kz:.2f}, 有{'/'.join(peak_desc)})"

    # 第2类：有波峰但不符合规律
    if peak_count >= 1:
        reasons = []
        if not (has_morning_peak or has_evening_peak):
            reasons.append(f"高峰时间异常({peak_times})")
        if not (1.3 <= kz <= 2.5):
            reasons.append(f"Kz={kz:.2f}超出正常范围")
        if not night_normal:
            reasons.append(f"夜间流量偏高(夜/日比={night_day_ratio:.2f})")
        return 2, "; ".join(reasons) if reasons else "有波峰但不符合典型规律"

    return 2, "特征不明确"


def _build_description(features: dict, category: int, peak_periods: list[str], valley_periods: list[str]) -> str:
    """生成排污规律描述"""
    if category == 1:
        desc = "流量特征曲线呈现明显的波峰和波谷特征，符合居民生活污水排放规律。"
        if peak_periods:
            desc += f"波峰主要出现在{'、'.join(peak_periods[:3])}，与早晚用水高峰期相吻合。"
        if valley_periods:
            desc += f"波谷主要出现在{'、'.join(valley_periods[:2])}，与夜间和午间用水低峰期一致。"
    elif category == 2:
        desc = "流量特征曲线有波峰或波谷，但不符合典型的生活污水排放规律。"
        if peak_periods:
            desc += f"波峰主要出现在{'、'.join(peak_periods[:3])}。"
        if valley_periods:
            desc += f"波谷主要出现在{'、'.join(valley_periods[:2])}。"
        desc += "可能受下游泵站启停、管道顶托或其他因素影响。"
    else:
        desc = "流量特征曲线无明显波峰波谷特征，曲线较为平坦，建议进一步排查管道运行状况。"

    return desc


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_pattern_analysis(
    dry_curve_data: dict[str, pd.DataFrame],
    combined_xlsx: Path,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行排污规律分析

    Args:
        dry_curve_data: 旱天特征曲线数据（从内存传入）
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        config: 可选配置参数

    Returns:
        {
            "pattern_df": pd.DataFrame,    # 分析结果
            "descriptions": dict,           # 点位描述
        }
    """
    # 合并配置
    cfg = PatternConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    rows: list[dict] = []
    descriptions: dict[str, str] = {}

    for point_name, curve in dry_curve_data.items():
        # 计算特征
        features = _calculate_features(curve)

        # 提取波峰波谷时段
        peak_periods, valley_periods = _extract_peak_valley_periods(curve)

        # 分类
        category, reason = _classify_pattern(features)

        # 生成描述
        description = _build_description(features, category, peak_periods, valley_periods)

        # 分类名称
        category_names = {
            1: "第1类-符合生活用水规律",
            2: "第2类-有波峰但不符合典型规律",
            3: "第3类-曲线平坦/异常",
        }

        row = {
            "点位编号": point_name,
            "分类": category,
            "分类名称": category_names.get(category, "未分类"),
            "Kz值": round(features["kz"], 2),
            "峰谷比": round(features["peak_valley_ratio"], 2) if features["peak_valley_ratio"] < 99 else "N/A",
            "峰数量": features["peak_count"],
            "波峰时段": "、".join(peak_periods[:3]) if peak_periods else "",
            "波谷时段": "、".join(valley_periods[:2]) if valley_periods else "",
            "诊断理由": reason,
            "排污规律描述": description,
        }
        rows.append(row)
        descriptions[point_name] = description

    # 创建结果 DataFrame
    pattern_df = pd.DataFrame(rows)

    # 输出到综合分析结果.xlsx
    _save_to_excel(
        pattern_df,
        combined_xlsx,
        "排污规律分析",
        ["点位编号", "分类", "分类名称", "Kz值", "峰谷比", "峰数量", "波峰时段", "波谷时段", "诊断理由", "排污规律描述"]
    )
    print(f"保存排污规律分析: {combined_xlsx}")

    # 统计
    cat_counts = pattern_df["分类"].value_counts().to_dict()
    print(f"\n排污规律分析完成:")
    print(f"  - 第1类(符合生活规律): {cat_counts.get(1, 0)} 个点位")
    print(f"  - 第2类(不符合典型规律): {cat_counts.get(2, 0)} 个点位")
    print(f"  - 第3类(曲线平坦/异常): {cat_counts.get(3, 0)} 个点位")

    return {
        "pattern_df": pattern_df,
        "descriptions": descriptions,
    }
