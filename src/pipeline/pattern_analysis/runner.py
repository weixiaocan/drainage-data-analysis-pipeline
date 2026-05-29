"""排污规律分析模块入口

统一接口: run(config: Config, logger, dry_curve_data=None) -> dict

输出:
    - config.combined_xlsx_path 的 "排污规律分析" Sheet
    - 返回值: {pattern_df, descriptions}
"""

import logging
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from src.core.config import Config
from openpyxl import load_workbook

from .analyzer import run_pattern_analysis


def _load_dry_curve_data_from_excel(combined_xlsx: Path, logger: logging.Logger) -> dict[str, pd.DataFrame]:
    """从 Excel 读取旱天特征曲线数据

    Args:
        combined_xlsx: 综合分析结果 Excel 文件

    Returns:
        {点位编号: 特征曲线 DataFrame}
    """
    dry_curve_data: dict[str, pd.DataFrame] = {}

    try:
        wb = load_workbook(combined_xlsx, data_only=True)

        # 查找旱天特征曲线 sheet（每个点位一个 sheet）
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("特征曲线_"):
                ws = wb[sheet_name]
                point_name = sheet_name.replace("特征曲线_", "")

                # 读取数据
                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
                    if row[0] is not None:
                        data.append(row)

                if data:
                    df = pd.DataFrame(data, columns=["时间", "流量(L/s)", "液位(m)", "流速(m/s)"])
                    df = df.dropna(subset=["时间"])
                    # 重新设置时间索引
                    df["时间"] = pd.date_range("00:00:00", "23:59:00", freq="T")[:len(df)]
                    df = df.set_index("时间")
                    df = df.rename(columns={"流量(L/s)": "f", "液位(m)": "l", "流速(m/s)": "velo"})
                    dry_curve_data[point_name] = df

        wb.close()

    except Exception as e:
        logger.warning(f"读取旱天特征曲线数据失败: {e}")

    return dry_curve_data


def run(
    config: Config,
    logger: logging.Logger,
    dry_curve_data: dict[str, pd.DataFrame] | None = None,
) -> dict[str, Any]:
    """
    排污规律分析入口。

    输入:
        - 旱天特征曲线数据（从内存传入，或从 Excel 读取）

    输出:
        - config.combined_xlsx_path 的 "排污规律分析" Sheet

    返回:
        {
            "pattern_df": pd.DataFrame,    # 分析结果
            "descriptions": dict,           # 点位描述
        }
    """
    combined_xlsx = config.combined_xlsx_path

    logger.info(f"开始排污规律分析")
    logger.info(f"  综合分析结果: {combined_xlsx}")

    # 如果没有传入 dry_curve_data，从 Excel 读取
    if dry_curve_data is None:
        logger.info("  从 Excel 读取旱天特征曲线数据...")
        dry_curve_data = _load_dry_curve_data_from_excel(combined_xlsx, logger)

    if not dry_curve_data:
        logger.warning("未找到旱天特征曲线数据，跳过排污规律分析")
        return {
            "pattern_df": pd.DataFrame(),
            "descriptions": {},
        }

    logger.info(f"  加载点位数: {len(dry_curve_data)}")

    # 执行分析
    result = run_pattern_analysis(
        dry_curve_data=dry_curve_data,
        combined_xlsx=combined_xlsx,
        config=None,  # 使用默认配置
    )

    # 统计
    pattern_df = result["pattern_df"]
    if not pattern_df.empty:
        cat_counts = pattern_df["分类"].value_counts().to_dict()
        logger.info(f"排污规律分析完成")
        logger.info(f"  第1类(符合生活规律): {cat_counts.get(1, 0)} 个点位")
        logger.info(f"  第2类(不符合典型规律): {cat_counts.get(2, 0)} 个点位")
        logger.info(f"  第3类(曲线平坦/异常): {cat_counts.get(3, 0)} 个点位")

    return result
