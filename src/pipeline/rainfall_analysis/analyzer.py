"""降雨分析核心逻辑

从 colleague_tool/code/read_prepro_rain_data.py 和 analyze_rain.py 提取核心逻辑。
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class RainfallConfig:
    """降雨分析配置参数"""
    min_interval: float = 12.0    # 场次降雨划分时间间隔（小时）
    min_rainfall: float = 1.0     # 最小降雨量阈值（mm）


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV"""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as err:
            last_err = err
    if last_err:
        raise last_err
    raise RuntimeError(f"无法读取 CSV: {path}")


def _load_rain_data(rainfall_file: Path) -> pd.DataFrame:
    """加载降雨数据并预处理

    自动检测数据频率（分钟级/小时级/日级），统一输出为分钟级数据。
    小时级数据会展开为分钟级，均匀分配降雨量。
    """
    df = _read_csv_with_fallback(rainfall_file)

    # 检测时间列和雨量列
    cols = [str(c).strip().lower() for c in df.columns]
    time_col = None
    rain_col = None

    for i, c in enumerate(df.columns):
        c_lower = str(c).strip().lower()
        if 't' == c_lower or 'time' in c_lower:
            time_col = df.columns[i]
        if 'rain' in c_lower:
            rain_col = df.columns[i]

    if time_col is None:
        time_col = df.columns[0]
    if rain_col is None:
        rain_col = df.columns[1]

    # 解析时间
    df[time_col] = pd.to_datetime(df[time_col], errors="coerce")
    df = df.dropna(subset=[time_col]).copy()

    # 转换雨量
    df[rain_col] = pd.to_numeric(df[rain_col], errors="coerce").fillna(0.0)

    # 重命名列
    df = df.rename(columns={time_col: "time", rain_col: "rain"})

    # 检测数据频率
    df_sorted = df.sort_values("time")
    time_diffs = df_sorted["time"].diff().dropna()
    median_diff = time_diffs.median()

    if median_diff <= pd.Timedelta(minutes=5):
        # 分钟级数据，直接填充连续时间序列
        freq = "minute"
        print(f"检测到分钟级降雨数据（间隔 {median_diff}）")
        time_start = df["time"].min()
        time_end = df["time"].max()
        full_index = pd.date_range(time_start, time_end, freq="T")
        full_df = pd.DataFrame({"time": full_index})
        df = full_df.merge(df, on="time", how="left")
        df["rain"] = df["rain"].fillna(0.0)

    elif median_diff <= pd.Timedelta(hours=3):
        # 小时级数据，展开为分钟级并均匀分配
        freq = "hourly"
        print(f"检测到小时级降雨数据（间隔 {median_diff}），展开为分钟级数据")
        df = _expand_hourly_to_minute(df_sorted)
    else:
        # 日级或其他频率，暂不支持
        raise ValueError(f"不支持的降雨数据频率: {median_diff}，请提供分钟级或小时级数据")

    return df.set_index("time")


def _expand_hourly_to_minute(df: pd.DataFrame) -> pd.DataFrame:
    """将小时级降雨数据展开为分钟级，均匀分配降雨量

    例如：1:00 的 10mm 降雨量 -> 0:00-0:59 每分钟 10/60 ≈ 0.167mm

    Args:
        df: 包含 'time' 和 'rain' 列的 DataFrame，时间为整点小时

    Returns:
        展开后的分钟级 DataFrame
    """
    records = []
    for _, row in df.iterrows():
        hour_start = row["time"]
        rain_per_minute = row["rain"] / 60.0

        # 生成该小时的60分钟记录
        for minute in range(60):
            minute_time = hour_start + pd.Timedelta(minutes=minute)
            records.append({
                "time": minute_time,
                "rain": rain_per_minute,
            })

    result = pd.DataFrame(records)

    # 填充可能缺失的小时（确保连续）
    time_start = result["time"].min()
    time_end = result["time"].max()
    full_index = pd.date_range(time_start, time_end, freq="T")
    full_df = pd.DataFrame({"time": full_index})
    result = full_df.merge(result, on="time", how="left")
    result["rain"] = result["rain"].fillna(0.0)

    return result


def _get_daily_rain(rain_data: pd.DataFrame) -> pd.DataFrame:
    """计算日降雨量"""
    daily = rain_data.resample("D").sum()
    daily = daily.reset_index()
    daily.columns = ["日期", "日降雨量(mm)"]
    return daily


def _time_split(df: pd.DataFrame, n_hours: float) -> list[tuple[datetime, datetime]]:
    """场次降雨划分

    Args:
        df: 降雨数据 DataFrame，index 为时间，包含 'rain' 列
        n_hours: 划分时间间隔（小时）

    Returns:
        [(start_time, end_time), ...] 场次降雨起止时间列表
    """
    # 去掉 0 值
    df_nonzero = df[df["rain"] > 0].copy()
    if df_nonzero.empty:
        return []

    timestamps = df_nonzero.index
    rain_rng: list[tuple[datetime, datetime]] = []
    time_nodes = [timestamps[0]]

    for i in range(1, len(timestamps)):
        diff = (timestamps[i] - timestamps[i - 1]).total_seconds()
        if diff >= n_hours * 3600:
            rain_rng.append((time_nodes[-1], timestamps[i - 1]))
            time_nodes.append(timestamps[i])

    rain_rng.append((time_nodes[-1], timestamps[-1]))
    return rain_rng


def _get_rain_info(
    rain_rng: list[tuple[datetime, datetime]],
    rain_data: pd.DataFrame,
    min_rain: float,
) -> pd.DataFrame:
    """提取场次降雨特征值

    Args:
        rain_rng: 场次降雨起止时间列表
        rain_data: 降雨数据 DataFrame
        min_rain: 最小降雨量阈值（mm）

    Returns:
        场次降雨统计 DataFrame
    """
    records: list[dict] = []

    for i, (start, end) in enumerate(rain_rng):
        event_data = rain_data.loc[start:end]
        total_rain = event_data["rain"].sum()

        if total_rain > min_rain:
            duration = (end - start).total_seconds() / 3600  # 小时
            records.append({
                "场次编号": i + 1,
                "开始时间": start,
                "结束时间": end,
                "总降雨量(mm)": round(total_rain, 2),
                "降雨历时(h)": round(duration, 2),
                "最大1分钟降雨量(mm)": round(event_data["rain"].max(), 2),
                "最大5分钟降雨量(mm)": round(event_data["rain"].rolling(5).sum().max(), 2),
                "最大10分钟降雨量(mm)": round(event_data["rain"].rolling(10).sum().max(), 2),
                "最大60分钟降雨量(mm)": round(event_data["rain"].rolling(60).sum().max(), 2),
                "最大24小时降雨量(mm)": round(event_data["rain"].rolling(1440).sum().max(), 2),
                "平均强度(mm/h)": round(total_rain / duration, 2) if duration > 0 else 0,
            })

    return pd.DataFrame(records)


def _classify_rain_level(total_rain: float) -> str:
    """根据总降雨量判断降雨等级"""
    if total_rain < 10:
        return "小雨"
    elif total_rain < 25:
        return "中雨"
    elif total_rain < 50:
        return "大雨"
    elif total_rain < 100:
        return "暴雨"
    elif total_rain < 250:
        return "大暴雨"
    else:
        return "特大暴雨"


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # 打开或创建工作簿
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 删除已存在的 sheet
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # 创建新 sheet
    ws = wb.create_sheet(sheet_name)

    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 替换表头
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    # 格式化
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_rainfall_analysis(
    rainfall_file: Path,
    combined_xlsx: Path,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行降雨分析

    Args:
        rainfall_file: 降雨数据 CSV 文件
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        config: 可选配置参数

    Returns:
        {
            "daily_rain": pd.DataFrame,      # 日降雨量统计
            "event_rain": pd.DataFrame,       # 场次降雨统计
            "rain_data": pd.DataFrame,        # 预处理后的降雨数据
            "event_data_dict": dict,          # 场次降雨详细数据（供后续模块使用）
        }
    """
    # 合并配置
    cfg = RainfallConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载降雨数据
    print(f"读取降雨数据: {rainfall_file}")
    rain_data = _load_rain_data(rainfall_file)
    print(f"  - 时间范围: {rain_data.index.min()} ~ {rain_data.index.max()}")
    print(f"  - 总降雨量: {rain_data['rain'].sum():.2f} mm")

    # 日降雨量统计
    print("计算日降雨量统计")
    daily_rain = _get_daily_rain(rain_data)
    rainy_days = (daily_rain["日降雨量(mm)"] > 0).sum()
    print(f"  - 降雨日数: {rainy_days}")

    # 场次降雨划分
    print(f"场次降雨划分: 间隔 {cfg.min_interval} 小时, 最小降雨量 {cfg.min_rainfall} mm")
    rain_rng = _time_split(rain_data, cfg.min_interval)
    event_rain = _get_rain_info(rain_rng, rain_data, cfg.min_rainfall)
    print(f"  - 场次降雨数: {len(event_rain)}")

    # 添加降雨等级
    if not event_rain.empty:
        event_rain["降雨等级"] = event_rain["总降雨量(mm)"].apply(_classify_rain_level)

    # 输出到综合分析结果.xlsx
    # 日降雨量 sheet
    _save_to_excel(
        daily_rain,
        combined_xlsx,
        "日降雨量统计",
        ["日期", "日降雨量(mm)"]
    )
    print(f"保存日降雨量统计: {combined_xlsx}")

    # 场次降雨 sheet
    _save_to_excel(
        event_rain,
        combined_xlsx,
        "场次降雨统计",
        ["场次编号", "开始时间", "结束时间", "总降雨量(mm)", "降雨历时(h)",
         "最大1分钟降雨量(mm)", "最大5分钟降雨量(mm)", "最大10分钟降雨量(mm)",
         "最大60分钟降雨量(mm)", "最大24小时降雨量(mm)", "平均强度(mm/h)", "降雨等级"]
    )
    print(f"保存场次降雨统计: {combined_xlsx}")

    # 构建场次降雨数据字典（供后续模块在内存中使用）
    event_data_dict: dict[int, dict] = {}
    for _, row in event_rain.iterrows():
        event_id = int(row["场次编号"])
        start = row["开始时间"]
        end = row["结束时间"]
        event_data_dict[event_id] = {
            "start": start,
            "end": end,
            "total_rain": row["总降雨量(mm)"],
            "duration": row["降雨历时(h)"],
            "rain_level": row["降雨等级"],
            "data": rain_data.loc[start:end].copy(),
        }

    return {
        "daily_rain": daily_rain,
        "event_rain": event_rain,
        "rain_data": rain_data,
        "event_data_dict": event_data_dict,
    }
