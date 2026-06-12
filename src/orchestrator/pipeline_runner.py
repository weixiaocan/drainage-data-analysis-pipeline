"""
orchestrator.pipeline_runner - Pipeline 流程编排核心

负责串联所有 pipeline 模块，实现完整的分析流程。

主要组件:
- ModuleInfo: 模块元信息
- PipelineState: Pipeline 执行状态
- Orchestrator: 流程编排器
"""

from __future__ import annotations

import importlib
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd

from src.core.config import Config


@dataclass
class ModuleInfo:
    """模块元信息"""

    name: str  # 模块名称
    runner_path: str  # runner 模块路径，如 "pipeline.data_filter.runner"
    critical: bool = True  # 是否为核心模块（核心模块失败终止流程）

    # 上游数据依赖
    needs_dry_curve_data: bool = False
    needs_event_data: bool = False
    needs_rain_data: bool = False  # 需要降雨数据（用于绑制 RDII 曲线）

    # 条件执行
    condition: str | None = None  # "has_rainfall_data" 或 None


@dataclass
class PipelineState:
    """Pipeline 执行状态"""

    # 已执行的模块结果 {module_name: result_dict}
    data: dict[str, dict[str, Any]] = field(default_factory=dict)

    # 已执行的模块列表（用于复用判断）
    executed_modules: list[str] = field(default_factory=list)

    # 失败的模块列表
    failed_modules: list[str] = field(default_factory=list)

    # 跳过的模块列表（复用）
    skipped_modules: list[str] = field(default_factory=list)

    # 条件执行状态
    has_rainfall_data: bool = True  # 是否有降雨数据
    rainfall_check_done: bool = False  # 是否已检查降雨数据

    def get_dry_curve_data(self) -> dict[str, pd.DataFrame] | None:
        """获取旱天特征曲线数据（从 dry_analysis 模块）"""
        if "dry_analysis" in self.data:
            return self.data["dry_analysis"].get("dry_curve_data")
        return None

    def get_dry_curve_data_workday(self) -> dict[str, pd.DataFrame] | None:
        """获取工作日旱天特征曲线数据（从 dry_analysis 模块）"""
        if "dry_analysis" in self.data:
            return self.data["dry_analysis"].get("dry_curve_data_workday")
        return None

    def get_dry_curve_data_weekend(self) -> dict[str, pd.DataFrame] | None:
        """获取周末旱天特征曲线数据（从 dry_analysis 模块）"""
        if "dry_analysis" in self.data:
            return self.data["dry_analysis"].get("dry_curve_data_weekend")
        return None

    def get_day_num(self) -> pd.DataFrame | None:
        """获取工作日/周末天数统计（从 dry_analysis 模块）"""
        if "dry_analysis" in self.data:
            return self.data["dry_analysis"].get("day_num")
        return None

    def get_event_data(self) -> dict[int, dict] | None:
        """获取场次降雨数据（从 rainfall_analysis 模块）"""
        if "rainfall_analysis" in self.data:
            return self.data["rainfall_analysis"].get("event_data_dict")
        return None

    def get_rain_data(self) -> pd.DataFrame | None:
        """获取降雨数据（从 rainfall_analysis 模块）"""
        if "rainfall_analysis" in self.data:
            return self.data["rainfall_analysis"].get("rain_data")
        return None


class Orchestrator:
    """
    Pipeline 流程编排器

    负责:
    1. 按顺序执行所有模块
    2. 在介入点暂停等待用户确认
    3. 传递上游数据给下游模块
    4. 处理模块失败（核心模块终止，辅助模块跳过）
    5. 支持复用已有输出
    """

    # 模块注册表（执行顺序）
    MODULES: list[ModuleInfo] = [
        # === 公共前置 ===
        ModuleInfo("data_stats", "src.pipeline.data_stats.runner", critical=False),
        ModuleInfo("data_filter", "src.pipeline.data_filter.runner", critical=True),
        # === 介入点 1 ===
        # === 雨天路径（条件执行）===
        ModuleInfo("rainfall_analysis", "src.pipeline.rainfall_analysis.runner", critical=False, condition="has_rainfall_data"),
        # === 介入点 2（条件触发）===
        # === 旱天路径 ===
        ModuleInfo("dry_analysis", "src.pipeline.dry_analysis.runner", critical=True),
        ModuleInfo("event_stats", "src.pipeline.event_stats.runner", critical=False, needs_event_data=True, condition="has_rainfall_data"),
        ModuleInfo("pattern_analysis", "src.pipeline.pattern_analysis.runner", critical=True, needs_dry_curve_data=True),
        # === 介入点 3 ===
        ModuleInfo("rdii_analysis", "src.pipeline.rdii_analysis.runner", critical=False, needs_dry_curve_data=True, needs_event_data=True, needs_rain_data=True, condition="has_rainfall_data"),
        ModuleInfo("risk_analysis", "src.pipeline.risk_analysis.runner", critical=True, needs_event_data=True),
        ModuleInfo("report_assembler", "src.pipeline.report_assembler.runner", critical=True, needs_dry_curve_data=True),
    ]

    # 介入点定义
    INTERVENTION_POINTS: list[dict[str, Any]] = [
        {
            "after": "data_filter",
            "message": "\n" + "=" * 60 + "\n【介入点 1】数据筛选完成\n请审核筛选结果.xlsx，确认数据质量后继续。\n" + "=" * 60,
            "type": "review",
        },
        {
            "after": "rainfall_analysis",
            "message": "\n" + "=" * 60 + "\n【介入点 2】降雨分析完成\n请在 baseinfo.xlsx 的「降雨场次选择」Sheet 中填写选中的场次编号。\n填写完成后输入 y 继续。\n" + "=" * 60,
            "type": "reload",
            "condition": "has_rainfall_data",  # 仅在有降雨数据时触发
        },
        {
            "after": "pattern_analysis",
            "message": "\n" + "=" * 60 + "\n【介入点 3】排污规律分析完成\n请审核综合分析结果.xlsx 的「排污规律分析」Sheet。\n确认后输入 y 继续。\n" + "=" * 60,
            "type": "review",
        },
    ]

    def __init__(self, config: Config, logger=None):
        """
        初始化编排器。

        参数:
            config: 全局配置对象
            logger: 日志器（可选，默认使用 print）
        """
        self.config = config
        self.logger = logger
        self.state = PipelineState()
        self.non_interactive = False  # 非交互模式（跳过所有确认）

    def _log(self, level: str, msg: str) -> None:
        """输出日志"""
        if self.logger:
            getattr(self.logger, level)(msg)
        else:
            print(f"[{level.upper()}] {msg}")

    def _check_rainfall_data(self) -> bool:
        """
        检查降雨数据是否存在且有效。

        Returns:
            True: 降雨数据存在且有效
            False: 降雨数据不存在或为空
        """
        rainfall_path = self.config.rainfall_data_path

        # 检查文件是否存在
        if not rainfall_path.exists():
            self._log("info", f"降雨数据文件不存在: {rainfall_path}")
            return False

        # 检查文件是否为空
        if rainfall_path.stat().st_size == 0:
            self._log("info", "降雨数据文件为空")
            return False

        # 尝试读取文件验证有效性
        try:
            import pandas as pd
            df = pd.read_csv(rainfall_path, nrows=5)
            if df.empty or len(df.columns) == 0:
                self._log("info", "降雨数据文件无有效数据")
                return False
            # 检查是否有数据列
            if len(df) == 0:
                self._log("info", "降雨数据文件无数据行")
                return False
        except Exception as e:
            self._log("warning", f"读取降雨数据失败: {e}")
            return False

        self._log("info", f"检测到有效降雨数据: {rainfall_path}")
        return True

    def _should_trigger_intervention(self, point: dict[str, Any]) -> bool:
        """
        检查是否应该触发介入点。

        Args:
            point: 介入点配置

        Returns:
            True: 应该触发介入点
            False: 跳过介入点
        """
        condition = point.get("condition")
        if condition == "has_rainfall_data":
            return self.state.has_rainfall_data
        return True  # 默认触发

    def run(self, stop_before: str | None = None) -> bool:
        """
        执行完整的 Pipeline 流程。

        参数:
            stop_before: 停止在指定模块之前（不执行该模块），如 "report_assembler"

        返回:
            True: 所有模块执行成功
            False: 有核心模块失败
        """
        self._log("info", "=" * 60)
        self._log("info", "开始执行 Pipeline")
        if stop_before:
            self._log("info", f"将在 [{stop_before}] 模块之前停止")
        self._log("info", "=" * 60)

        # 构建介入点索引
        intervention_after = {p["after"]: p for p in self.INTERVENTION_POINTS}

        success = True

        for module_info in self.MODULES:
            # 检查是否停止在指定模块之前
            if stop_before and module_info.name == stop_before:
                self._log("info", f"到达停止点 [{module_info.name}]，跳过后续所有模块")
                break

            # 条件执行检查
            if module_info.condition == "has_rainfall_data":
                if not self.state.rainfall_check_done:
                    self.state.has_rainfall_data = self._check_rainfall_data()
                    self.state.rainfall_check_done = True
                    if not self.state.has_rainfall_data:
                        self._log("info", "未检测到有效降雨数据，将跳过雨天分析路径")

                if not self.state.has_rainfall_data:
                    self._log("info", f"跳过模块 [{module_info.name}]（无降雨数据）")
                    self.state.skipped_modules.append(module_info.name)
                    continue

            # 检查是否跳过该模块（复用）
            if self._should_skip_module(module_info):
                self._log("info", f"跳过模块 [{module_info.name}]（输出已存在，用户选择复用）")
                self.state.skipped_modules.append(module_info.name)
                # 即使跳过，也需要从文件加载数据到 state
                self._load_module_output_to_state(module_info)
                continue

            # 执行模块
            result = self._run_module(module_info)

            if result is None:
                # 模块失败
                self.state.failed_modules.append(module_info.name)
                if module_info.critical:
                    self._log("error", f"核心模块 [{module_info.name}] 失败，终止 Pipeline")
                    success = False
                    break
                else:
                    self._log("warning", f"辅助模块 [{module_info.name}] 失败，跳过继续")
                    # 创建占位 Sheet
                    self._handle_auxiliary_module_failure(module_info.name)
                    continue

            # 保存结果
            self.state.data[module_info.name] = result
            self.state.executed_modules.append(module_info.name)

            # 检查是否有介入点
            if module_info.name in intervention_after:
                point = intervention_after[module_info.name]
                if self._should_trigger_intervention(point):
                    if not self._wait_for_user(point):
                        # 用户选择退出
                        self._log("info", "用户选择退出 Pipeline")
                        return False

        self._log("info", "=" * 60)
        self._log("info", "Pipeline 执行完成")
        self._log("info", f"  成功模块: {len(self.state.executed_modules)}")
        self._log("info", f"  跳过模块: {len(self.state.skipped_modules)}")
        self._log("info", f"  失败模块: {len(self.state.failed_modules)}")
        self._log("info", f"  降雨数据: {'有' if self.state.has_rainfall_data else '无'}")
        self._log("info", "=" * 60)

        return success

    def _run_module(self, module_info: ModuleInfo) -> dict[str, Any] | None:
        """
        执行单个模块。

        返回:
            模块结果字典，失败返回 None
        """
        self._log("info", f"开始执行模块 [{module_info.name}]")

        try:
            # 动态导入 runner
            runner_module = importlib.import_module(module_info.runner_path)
            run_func: Callable = runner_module.run

            # 准备参数
            kwargs = self._prepare_module_kwargs(module_info)

            # 执行
            result = run_func(self.config, self.logger, **kwargs)

            self._log("info", f"模块 [{module_info.name}] 执行成功")
            return result

        except Exception as e:
            self._log("error", f"模块 [{module_info.name}] 执行失败: {e}")
            return None

    def _prepare_module_kwargs(self, module_info: ModuleInfo) -> dict[str, Any]:
        """
        准备模块参数（上游数据传递）
        """
        kwargs: dict[str, Any] = {}

        if module_info.needs_dry_curve_data:
            dry_curve_data = self.state.get_dry_curve_data()
            if dry_curve_data:
                kwargs["dry_curve_data"] = dry_curve_data
                self._log("info", f"  传递上游数据: dry_curve_data ({len(dry_curve_data)} 个点位)")

            # 为 pattern_analysis 传递完整数据
            if module_info.name == "pattern_analysis":
                dry_curve_data_workday = self.state.get_dry_curve_data_workday()
                dry_curve_data_weekend = self.state.get_dry_curve_data_weekend()
                day_num = self.state.get_day_num()

                if dry_curve_data_workday:
                    kwargs["dry_curve_data_workday"] = dry_curve_data_workday
                if dry_curve_data_weekend:
                    kwargs["dry_curve_data_weekend"] = dry_curve_data_weekend
                if day_num is not None:
                    kwargs["day_num"] = day_num

        if module_info.needs_event_data:
            event_data = self.state.get_event_data()
            if event_data:
                kwargs["event_data"] = event_data
                self._log("info", f"  传递上游数据: event_data ({len(event_data)} 个场次)")

        if module_info.needs_rain_data:
            rain_data = self.state.get_rain_data()
            if rain_data is not None:
                kwargs["rain_data"] = rain_data
                self._log("info", f"  传递上游数据: rain_data ({len(rain_data)} 条记录)")

        # 为 risk_analysis 和 report_assembler 传递 has_rainfall_data 参数
        if module_info.name in ("risk_analysis", "report_assembler"):
            kwargs["has_rainfall_data"] = self.state.has_rainfall_data
            self._log("info", f"  传递参数: has_rainfall_data={self.state.has_rainfall_data}")

        return kwargs

    def _wait_for_user(self, point: dict[str, Any]) -> bool:
        """
        在介入点等待用户确认。

        返回:
            True: 用户确认继续
            False: 用户选择退出
        """
        print(point["message"])

        # 非交互模式下自动继续
        if self.non_interactive:
            self._log("info", "非交互模式，自动继续")
            # 如果是介入点 2，需要重新加载 baseinfo
            if point["type"] == "reload":
                self.config.reload_baseinfo()
                selected = self.config.selected_rainfall_events
                self._log("info", f"已重新加载 baseinfo.xlsx，选中场次: {selected}")
            return True

        while True:
            ans = input("输入 y 继续 (q 退出): ").strip().lower()
            if ans == "q":
                return False
            if ans == "y":
                # 如果是介入点 2，需要重新加载 baseinfo
                if point["type"] == "reload":
                    self.config.reload_baseinfo()
                    selected = self.config.selected_rainfall_events
                    print(f"已重新加载 baseinfo.xlsx，选中场次: {selected}")
                return True
            print("无效输入，请输入 y 或 q")

    def _should_skip_module(self, module_info: ModuleInfo) -> bool:
        """
        检查是否应该跳过模块（复用机制）

        检查输出文件是否存在，询问用户是否复用。
        """
        output_file = self._get_module_output_file(module_info)

        if output_file is None or not output_file.exists():
            return False

        # 非交互模式下总是重新执行
        if self.non_interactive:
            self._log("info", f"输出文件已存在，非交互模式下重新执行")
            return False

        # 输出文件已存在，询问用户
        print(f"\n模块 [{module_info.name}] 的输出文件已存在: {output_file}")
        ans = input("是否复用已有输出？(y 复用 / n 重新执行 / q 退出): ").strip().lower()

        if ans == "q":
            sys.exit(0)
        return ans == "y"

    def _get_module_output_file(self, module_info: ModuleInfo) -> Path | None:
        """获取模块的主要输出文件路径"""
        output_mapping = {
            "data_stats": self.config.combined_xlsx_path,
            "data_filter": self.config.filter_result_path,
            "rainfall_analysis": self.config.combined_xlsx_path,
            "dry_analysis": self.config.combined_xlsx_path,
            "event_stats": self.config.combined_xlsx_path,
            "pattern_analysis": self.config.combined_xlsx_path,
            "rdii_analysis": self.config.combined_xlsx_path,
            "risk_analysis": self.config.combined_xlsx_path,
            "report_assembler": self.config.report_output_path,
        }
        return output_mapping.get(module_info.name)

    def _load_module_output_to_state(self, module_info: ModuleInfo) -> None:
        """
        从输出文件加载模块结果到 state（用于复用场景）
        """
        # 对于复用的模块，我们需要从文件中读取数据到 state
        # 这样后续模块可以获取到上游数据
        combined_xlsx = self.config.combined_xlsx_path

        if module_info.name == "dry_analysis":
            # 从 Excel 读取旱天特征曲线数据
            dry_curve_data = self._load_dry_curve_data_from_excel(combined_xlsx)
            if dry_curve_data:
                self.state.data["dry_analysis"] = {"dry_curve_data": dry_curve_data}
                self._log("info", f"  从文件加载 dry_curve_data: {len(dry_curve_data)} 个点位")

        elif module_info.name == "rainfall_analysis":
            # 从 Excel 读取场次降雨数据
            event_data_dict = self._load_event_data_from_excel(combined_xlsx)
            if event_data_dict:
                self.state.data["rainfall_analysis"] = {"event_data_dict": event_data_dict}
                self._log("info", f"  从文件加载 event_data_dict: {len(event_data_dict)} 个场次")

    def _load_dry_curve_data_from_excel(self, combined_xlsx: Path) -> dict[str, pd.DataFrame]:
        """从 Excel 读取旱天特征曲线数据"""
        from openpyxl import load_workbook

        dry_curve_data: dict[str, pd.DataFrame] = {}

        try:
            wb = load_workbook(combined_xlsx, data_only=True)

            for sheet_name in wb.sheetnames:
                if sheet_name.startswith("特征曲线_"):
                    ws = wb[sheet_name]
                    point_name = sheet_name.replace("特征曲线_", "")

                    data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] is not None:
                            data.append(row)

                    if data:
                        df = pd.DataFrame(data, columns=["时间", "流量(L/s)", "液位(m)", "流速(m/s)"])
                        df = df.dropna(subset=["时间"])
                        df["时间"] = pd.date_range("00:00:00", "23:59:00", freq="min")[: len(df)]
                        df = df.set_index("时间")
                        df = df.rename(columns={"流量(L/s)": "f", "液位(m)": "l", "流速(m/s)": "velo"})
                        dry_curve_data[point_name] = df

            wb.close()

        except Exception as e:
            self._log("warning", f"读取旱天特征曲线数据失败: {e}")

        return dry_curve_data

    def _load_event_data_from_excel(self, combined_xlsx: Path) -> dict[int, dict]:
        """从 Excel 读取场次降雨数据"""
        from openpyxl import load_workbook

        event_data: dict[int, dict] = {}

        try:
            wb = load_workbook(combined_xlsx, data_only=True)

            if "场次降雨统计" in wb.sheetnames:
                ws = wb["场次降雨统计"]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue

                    event_id = int(row[0])
                    event_data[event_id] = {
                        "start": row[1],
                        "end": row[2],
                        "total_rain": float(row[3]) if row[3] else 0,
                        "duration": float(row[4]) if row[4] else 0,
                        "rain_level": row[11] if len(row) > 11 else "",
                    }

            wb.close()

        except Exception as e:
            self._log("warning", f"读取场次降雨数据失败: {e}")

        return event_data

    def _handle_auxiliary_module_failure(self, module_name: str) -> None:
        """
        处理辅助模块失败，创建占位 Sheet
        """
        combined_xlsx = self.config.combined_xlsx_path

        # 只有当文件存在时才写入
        if not combined_xlsx.exists():
            self._log("warning", f"综合分析结果文件不存在，跳过写入失败信息")
            return

        # 创建失败说明 DataFrame
        error_df = pd.DataFrame(
            [
                {
                    "状态": "计算失败",
                    "失败时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "说明": "该模块为辅助模块，失败后跳过。请检查数据或参数后重新运行该模块。",
                }
            ]
        )

        # 模块名到 Sheet 名映射
        sheet_mapping = {
            "rdii_analysis": "RDII分析",
        }

        sheet_name = sheet_mapping.get(module_name, module_name)

        try:
            with pd.ExcelWriter(combined_xlsx, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                error_df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._log("info", f"已在 [{sheet_name}] Sheet 中记录失败信息")
        except Exception as e:
            self._log("warning", f"写入失败信息到 Excel 时出错: {e}")
