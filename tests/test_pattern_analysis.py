"""测试 pattern_analysis 模块"""
from pathlib import Path

import pandas as pd
import pytest


PROJECT_ROOT = Path(__file__).parent.parent


class TestPatternAnalysis:
    """排污规律分析模块测试"""

    @pytest.fixture
    def dry_curve_file(self) -> Path:
        return PROJECT_ROOT / "outputs" / "旱天特征曲线.pickle"

    @pytest.fixture
    def output_dir(self, tmp_path: Path) -> Path:
        return tmp_path

    def test_run_pattern_analysis_import(self):
        """测试模块可导入"""
        from src.pipeline.pattern_analysis import run_pattern_analysis
        assert callable(run_pattern_analysis)

    def test_run_pattern_analysis_runs(
        self,
        dry_curve_file: Path,
        output_dir: Path,
    ):
        """测试运行不报错"""
        from src.pipeline.pattern_analysis import run_pattern_analysis

        result = run_pattern_analysis(
            dry_curve_file=dry_curve_file,
            output_dir=output_dir,
        )

        assert result is not None

    def test_run_pattern_analysis_returns_dict(
        self,
        dry_curve_file: Path,
        output_dir: Path,
    ):
        """测试返回结果格式"""
        from src.pipeline.pattern_analysis import run_pattern_analysis

        result = run_pattern_analysis(
            dry_curve_file=dry_curve_file,
            output_dir=output_dir,
        )

        assert "pattern_df" in result
        assert "descriptions" in result

        assert isinstance(result["pattern_df"], pd.DataFrame)
        assert isinstance(result["descriptions"], dict)

    def test_pattern_df_format(
        self,
        dry_curve_file: Path,
        output_dir: Path,
    ):
        """测试分析结果格式"""
        from src.pipeline.pattern_analysis import run_pattern_analysis

        result = run_pattern_analysis(
            dry_curve_file=dry_curve_file,
            output_dir=output_dir,
        )

        df = result["pattern_df"]

        required_cols = ["点位编号", "分类", "分类名称", "Kz值", "峰谷比", "排污规律描述"]
        for col in required_cols:
            assert col in df.columns

    def test_category_values(
        self,
        dry_curve_file: Path,
        output_dir: Path,
    ):
        """测试分类值范围"""
        from src.pipeline.pattern_analysis import run_pattern_analysis

        result = run_pattern_analysis(
            dry_curve_file=dry_curve_file,
            output_dir=output_dir,
        )

        df = result["pattern_df"]

        # 分类应该是 1, 2, 或 3
        for cat in df["分类"]:
            assert cat in [1, 2, 3]

    def test_output_xlsx_created(
        self,
        dry_curve_file: Path,
        output_dir: Path,
    ):
        """测试输出文件创建"""
        from src.pipeline.pattern_analysis import run_pattern_analysis
        from openpyxl import load_workbook

        run_pattern_analysis(
            dry_curve_file=dry_curve_file,
            output_dir=output_dir,
        )

        xlsx_file = output_dir / "分析结果.xlsx"
        assert xlsx_file.exists()

        wb = load_workbook(xlsx_file)
        assert "排污规律分析" in wb.sheetnames

    def test_descriptions_not_empty(
        self,
        dry_curve_file: Path,
        output_dir: Path,
    ):
        """测试描述不为空"""
        from src.pipeline.pattern_analysis import run_pattern_analysis

        result = run_pattern_analysis(
            dry_curve_file=dry_curve_file,
            output_dir=output_dir,
        )

        for point_name, desc in result["descriptions"].items():
            assert len(desc) > 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
