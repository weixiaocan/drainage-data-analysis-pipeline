"""测试 data_filter 模块"""
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook


# 项目根目录
PROJECT_ROOT = Path(__file__).parent.parent


class TestDataFilter:
    """数据筛选模块测试"""

    @pytest.fixture
    def flow_dir(self) -> Path:
        return PROJECT_ROOT / "data" / "flow"

    @pytest.fixture
    def rainfall_file(self) -> Path:
        return PROJECT_ROOT / "data" / "rainfall" / "降雨数据.csv"

    @pytest.fixture
    def output_file(self, tmp_path: Path) -> Path:
        return tmp_path / "筛选结果.xlsx"

    def test_run_data_filter_import(self):
        """测试模块可导入"""
        from src.pipeline.data_filter import run_data_filter
        assert callable(run_data_filter)

    def test_run_data_filter_generates_xlsx(
        self,
        flow_dir: Path,
        rainfall_file: Path,
        output_file: Path,
    ):
        """测试生成 xlsx 文件"""
        from src.pipeline.data_filter import run_data_filter

        result = run_data_filter(
            csv_dir=flow_dir,
            rainfall_file=rainfall_file,
            output_xlsx=output_file,
        )

        assert output_file.exists()

    def test_run_data_filter_returns_dict(
        self,
        flow_dir: Path,
        rainfall_file: Path,
        output_file: Path,
    ):
        """测试返回 dict 格式结果"""
        from src.pipeline.data_filter import run_data_filter

        result = run_data_filter(
            csv_dir=flow_dir,
            rainfall_file=rainfall_file,
            output_xlsx=output_file,
        )

        assert isinstance(result, dict)
        assert len(result) > 0

        # 检查格式
        for point_name, days in result.items():
            assert isinstance(point_name, str)
            assert isinstance(days, list)
            for day in days:
                assert isinstance(day, str)
                # 验证日期格式
                pd.to_datetime(day)  # 不抛异常即可

    def test_xlsx_has_green_fill(
        self,
        flow_dir: Path,
        rainfall_file: Path,
        output_file: Path,
    ):
        """测试 xlsx 中有绿色填充"""
        from src.pipeline.data_filter import run_data_filter

        result = run_data_filter(
            csv_dir=flow_dir,
            rainfall_file=rainfall_file,
            output_xlsx=output_file,
        )

        wb = load_workbook(output_file)
        ws = wb["筛选结果"]

        # 找到有有效旱天的点位
        has_green = False
        for row in range(3, ws.max_row + 1):  # 从第3行开始（跳过表头和雨量行）
            for col in range(2, ws.max_column):
                cell = ws.cell(row=row, column=col)
                fill = cell.fill
                if fill and fill.start_color:
                    color = str(fill.start_color.index).upper()
                    if color.endswith("92D050"):
                        has_green = True
                        break
            if has_green:
                break

        assert has_green, "xlsx 中应该有绿色填充的有效旱天单元格"

    def test_xlsx_format(
        self,
        flow_dir: Path,
        rainfall_file: Path,
        output_file: Path,
    ):
        """测试 xlsx 格式符合契约"""
        from src.pipeline.data_filter import run_data_filter

        result = run_data_filter(
            csv_dir=flow_dir,
            rainfall_file=rainfall_file,
            output_xlsx=output_file,
        )

        wb = load_workbook(output_file)
        ws = wb["筛选结果"]

        # 第1行表头
        assert ws.cell(row=1, column=1).value == "点位编号"

        # 第2行是雨量行
        assert ws.cell(row=2, column=1).value == "当天雨量"

        # 第3行起是点位数据
        assert ws.cell(row=3, column=1).value is not None


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
