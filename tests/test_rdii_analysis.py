"""测试 rdii_analysis 模块"""
from pathlib import Path
import pickle

import pandas as pd
import pytest


PROJECT_ROOT = Path(__file__).parent.parent


class TestRDIIAnalysis:
    """RDII分析模块测试"""

    @pytest.fixture
    def flow_dir(self) -> Path:
        return PROJECT_ROOT / "data" / "flow"

    @pytest.fixture
    def dry_curve_file(self) -> Path:
        return PROJECT_ROOT / "outputs" / "旱天特征曲线.pickle"

    @pytest.fixture
    def event_data_file(self) -> Path:
        return PROJECT_ROOT / "outputs" / "场次降雨数据.pickle"

    @pytest.fixture
    def output_dir(self, tmp_path: Path) -> Path:
        return tmp_path

    def test_run_rdii_analysis_import(self):
        """测试模块可导入"""
        from src.pipeline.rdii_analysis import run_rdii_analysis
        assert callable(run_rdii_analysis)

    def test_run_rdii_analysis_runs(
        self,
        flow_dir: Path,
        dry_curve_file: Path,
        event_data_file: Path,
        output_dir: Path,
    ):
        """测试运行不报错"""
        from src.pipeline.rdii_analysis import run_rdii_analysis

        result = run_rdii_analysis(
            flow_dir=flow_dir,
            dry_curve_file=dry_curve_file,
            event_data_file=event_data_file,
            output_dir=output_dir,
        )

        assert result is not None

    def test_run_rdii_analysis_returns_dict(
        self,
        flow_dir: Path,
        dry_curve_file: Path,
        event_data_file: Path,
        output_dir: Path,
    ):
        """测试返回结果格式"""
        from src.pipeline.rdii_analysis import run_rdii_analysis

        result = run_rdii_analysis(
            flow_dir=flow_dir,
            dry_curve_file=dry_curve_file,
            event_data_file=event_data_file,
            output_dir=output_dir,
        )

        assert "max_level" in result
        assert "avg_flow" in result
        assert "rdii_total" in result
        assert "overflow_total" in result
        assert "rdii_curve_data" in result

        assert isinstance(result["max_level"], pd.DataFrame)
        assert isinstance(result["avg_flow"], pd.DataFrame)
        assert isinstance(result["rdii_total"], pd.DataFrame)

    def test_output_files_created(
        self,
        flow_dir: Path,
        dry_curve_file: Path,
        event_data_file: Path,
        output_dir: Path,
    ):
        """测试输出文件创建"""
        from src.pipeline.rdii_analysis import run_rdii_analysis

        run_rdii_analysis(
            flow_dir=flow_dir,
            dry_curve_file=dry_curve_file,
            event_data_file=event_data_file,
            output_dir=output_dir,
        )

        # 检查 xlsx 文件
        xlsx_file = output_dir / "分析结果.xlsx"
        assert xlsx_file.exists()

        # 检查 pickle 文件
        pickle_file = output_dir / "RDII曲线数据.pickle"
        assert pickle_file.exists()

    def test_xlsx_sheets_created(
        self,
        flow_dir: Path,
        dry_curve_file: Path,
        event_data_file: Path,
        output_dir: Path,
    ):
        """测试 xlsx sheet 创建"""
        from src.pipeline.rdii_analysis import run_rdii_analysis
        from openpyxl import load_workbook

        run_rdii_analysis(
            flow_dir=flow_dir,
            dry_curve_file=dry_curve_file,
            event_data_file=event_data_file,
            output_dir=output_dir,
        )

        xlsx_file = output_dir / "分析结果.xlsx"
        wb = load_workbook(xlsx_file)

        expected_sheets = ["降雨事件最大液位", "降雨事件平均流量", "RDII总量统计", "雨天流量总量"]
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f"Sheet '{sheet}' not found"

    def test_config_parameters(
        self,
        flow_dir: Path,
        dry_curve_file: Path,
        event_data_file: Path,
        output_dir: Path,
    ):
        """测试配置参数"""
        from src.pipeline.rdii_analysis import run_rdii_analysis

        result = run_rdii_analysis(
            flow_dir=flow_dir,
            dry_curve_file=dry_curve_file,
            event_data_file=event_data_file,
            output_dir=output_dir,
            config={"rain_effect_delay": 24.0},
        )

        assert result is not None


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
